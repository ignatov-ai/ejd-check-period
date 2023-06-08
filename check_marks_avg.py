import itertools
import math
import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Fill, NamedStyle, Border
from math import ceil

def round_mark(num):
    num = int(num + (0.5 if num > 0 else -0.5))
    return num

def round_ap(num):
    if num < 2.6:
        return 2
    elif num < 3.6:
        return 3
    elif num < 4.6:
        return 4
    else:
        return 5

file_name_all_periods = '9ч-алг.xlsx'

book = openpyxl.load_workbook(filename='data/' + file_name_all_periods)

sheet_names = book.sheetnames

for s in range(len(sheet_names)):
    sheet = book.worksheets[s]

    # Проверяем, является ли ячейка A1 объединенной
    for row in sheet.iter_rows():
        # Инициализируем переменную для хранения предыдущей ячейки
        prev_cell = ''

        # Проходим по всем ячейкам в строке
        for cell in row:
            # Если текущая ячейка объединена с предыдущей, то разъединяем объединение и копируем данные из всех ячеек в новую ячейку
            if prev_cell and cell.coordinate == prev_cell.coordinate:
                range_string = prev_cell.coordinate + ':' + cell.coordinate
                sheet.unmerge_cells(range_string)

                new_cell = row[cell.column - 1]
                for merged_cell in sheet.merged_cells.ranges:
                    if merged_cell.min_row == prev_cell.row and merged_cell.max_row == prev_cell.row and merged_cell.min_col <= prev_cell.column and merged_cell.max_col >= cell.column:
                        for row_index in range(merged_cell.min_row, merged_cell.max_row + 1):
                            for col_index in range(merged_cell.min_col, merged_cell.max_col + 1):
                                source = sheet.cell(row=row_index, column=col_index)
                                target = sheet.cell(row=row_index, column=new_cell.column)
                                target.value = source.value

                # Обновляем переменную для хранения предыдущей ячейки
                prev_cell = new_cell
            else:
                # Обновляем переменную для хранения предыдущей ячейки
                prev_cell = cell

    for i in range(21, 24):
        for j in range(1,40,2):
            sheet.unmerge_cells(start_row=j, start_column=i, end_row=j+1, end_column=i)

    sheet.delete_cols(20, 5)

    i = 0

    # Создаем список букв
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    # Создаем комбинации из 1 и 2 символов
    col = itertools.chain(letters, itertools.product(letters, repeat=2))

    i = 0
    for s in col:
        t = ''.join(s)
        sheet.column_dimensions[t].width = 3
        if i == 400:
            break
        i += 1
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 30

    # Определяем количество заполненных строк в столбце A
    num_rows = len(list(sheet['A']))
    data_pages_count = num_rows // 50

    for i in range(1,data_pages_count):
        # Переносим область
        sheet.move_range('C' + str(1 + i*50) + ':S' + str((i+1)*50), rows=-(i*50), cols=(17 + (i-1)*17))

    # Удаляем перемещенные строки до конца таблицы
    sheet.delete_rows(41, sheet.max_row)

    # Меняем тип у оценок из str в int
    col = itertools.chain(letters, itertools.product(letters, repeat=2))
    for c in col:
        stop = False
        for row in range(2,40):
            cell = ''.join(c)+str(row)
            #print(cell,sheet[cell].value)
            if sheet[cell].value in ['1','2', '3', '4', '5']:
                sheet[cell].value = int(sheet[cell].value)
            # elif sheet[cell].value == None:
            #     print(cell)
            #     t = sheet[cell]
            #     t.value = ''

            if cell[:2] == 'OK':
                stop = True
                break
        if stop:
            break

    for i in range(4,sheet.max_row):
    #for i in range(4, 8):
        sum_period = 0
        count = 0
        ap_all = []
        for j in range(3,3 + 17 * data_pages_count - 1):
            mark = sheet.cell(row=i, column=j).value
            koef = sheet.cell(row=i, column=j+1).value

            if sheet.cell(row=3, column=j).value == 'оц' and (mark and koef) != '' \
                    and not isinstance(sheet.cell(row=i, column=j).value,str) \
                    and not isinstance(sheet.cell(row=i, column=j+1).value,str):
                print('оц =',mark,'коэф =',koef,'итог = ',mark*koef,sheet.cell(row=i, column=j))
                sum_period += mark*koef
                count += koef

            val = sheet.cell(row=2, column=j).value
            ap = ['АП1', 'АП2', 'АП3']
            if val in ap and count != 0:
                ap_mark = sheet.cell(row=i, column=j).value
                avg_mark = sum_period/count
                avg_mark_round = round_ap(avg_mark)
                print('Сумма =',sum_period,'| кол =', count,'| ср =', avg_mark,'| ср окр=', avg_mark_round,'| АП =',ap_mark)
                if ap_mark == avg_mark_round:
                    sheet.cell(row=i, column=j).fill = PatternFill('solid', fgColor='A6F16C')
                    sum_period = 0
                    count = 0
                    ap_all.append(ap_mark)
                else:
                    sheet.cell(row=i, column=j).fill = PatternFill('solid', fgColor='FF8B73')

            if val == 'Г' and len(ap_all) > 0:
                god = round_mark(sum(ap_all)/len(ap_all))
                if god == sheet.cell(row=i, column=j).value:
                    sheet.cell(row=i, column=j).fill = PatternFill('solid', fgColor='A6F16C')
                else:
                    sheet.cell(row=i, column=j).fill = PatternFill('solid', fgColor='FF8B73')


# Сохраняем книгу.
print('Файл создан')
book.save("example.xlsx")