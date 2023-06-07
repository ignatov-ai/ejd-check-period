import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Fill, NamedStyle, Border
from math import ceil

# Открытие файла с итоговыми оценками за периоды
file_name_all_periods = '10Ч-Костина-все'

book = openpyxl.load_workbook(filename='data/' + file_name_all_periods + '.xlsx')
sheet = book.worksheets[0]

data = []
for row in sheet.iter_rows(values_only=True):
    row_data = []
    for cell in row:
        if cell is not None:
            row_data.append(cell)
        else:
            row_data.append('')
    data.append(row_data)

for i in range (1,len(data[0])):
    if data[0][i] == '':
        data[0][i] = data[0][i-1]

del data[1]


for i in range(len(data)):
    for j in range(len(data[0])):
        print(data[i][j], end=';')
    print()
print()

headers = ['','Аттестационный период 1','Аттестационный период 2','Год']
i = 0
while i < len(data[1]):
    if data[1][i] not in headers:
        for j in range(len(data)):
            del data[j][i]
        i -= 1
    i += 1

math_col = data[0].index('Математика')
for j in range(len(data)):
    del data[j][math_col]

i = 0
while i < len(data[1])-1:
    if data[1][i] == 'Год' and data[1][i+1] != 'Аттестационный период 1':
        data[0].insert(i + 1,data[0][i+1])
        data[1].insert(i + 1, 'Аттестационный период 1')
        for j in range(2,len(data)):
            data[j].insert(i+1,'')
        i += 1
    i += 1

i = 0
while i < len(data[1])-1:
    if data[1][i] == 'Аттестационный период 1' and data[1][i+1] != 'Аттестационный период 2':
        data[0].insert(i + 1,data[0][i])
        data[1].insert(i + 1, 'Аттестационный период 2')
        for j in range(2,len(data)):
            data[j].insert(i+1,'')
        i += 1
    i += 1

for i in range(len(data)):
    for j in range(len(data[0])):
        print(data[i][j], end=';')
    print()

# for i in range(len(data[2])):
#     if data[2][i] == 'Год':
#         predmeti_p1.append(data[0][i])
#         if data[2][i-1] != 'Аттестационный период 2':
#             print('пусто')