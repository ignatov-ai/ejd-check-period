import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Fill, NamedStyle, Border
from math import ceil

# Укажите кол-во периодов для выгрузки
periods = 2

# просмотр содержимого папки data
folder = "data"
files = os.listdir(folder)

data_files = []
for s in files:
    if s.find('все') != -1:
        data_files.append(s)
print(data_files)

for file_name in data_files:
    book = openpyxl.load_workbook(filename='data/' + file_name)
    sheet = book.worksheets[0]

    data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = []
        for cell in row:
            if cell is not None:
                row_data.append(cell)
            else:
                row_data.append('')
        data.append(row_data)

    for i in range (1,len(data[0])):
        if data[0][i] == '':
            data[0][i] = data[0][i-1]

    del data[1]

    headers = ['','Аттестационный период 1','Аттестационный период 2','Год']
    i = 0
    while i < len(data[1]):
        if data[1][i] not in headers:
            for j in range(len(data)):
                del data[j][i]
            i -= 1
        i += 1

    math_col = data[0].index('Математика')
    for j in range(len(data)):
        del data[j][math_col]

    i = 0
    while i < len(data[1])-1:
        if data[1][i] == 'Год' and data[1][i+1] != 'Аттестационный период 1':
            data[0].insert(i + 1,data[0][i+1])
            data[1].insert(i + 1, 'Аттестационный период 1')
            for j in range(2,len(data)):
                data[j].insert(i+1,'')
            i += 1
        i += 1

    i = 0
    while i < len(data[1])-1:
        if data[1][i] == 'Аттестационный период 1' and data[1][i+1] != 'Аттестационный период 2':
            data[0].insert(i + 1,data[0][i])
            data[1].insert(i + 1, 'Аттестационный период 2')
            for j in range(2,len(data)):
                data[j].insert(i+1,'')
            i += 1
        i += 1

    predmeti = ';'.join(data[0])

    predmeti_p1 = ['']
    for i in range(len(data[1])):
        if data[1][i] == 'Год':
            predmeti_p1.append(data[0][i])

    # Трассировка вывода предметов
    # print(predmeti_p1)

    # создание книги для проверенной выгрузки
    out_book = openpyxl.Workbook()
    out_book.remove(out_book.active)
    out_sheet = out_book.create_sheet("Проверка ГОД")
    out_sheet = out_book.active
    out_sheet.column_dimensions['A'].width = 30

    for col in 'BCDEFGHIJKLMNOPQRSTUVWXWZ':
        out_sheet.column_dimensions[col].width = 5


    # Обработка таблицы с периодами
    def markToInt(x):
        marks = ['2', '3', '4', '5']
        if x in marks:
            return int(x)
        elif x == '':
            return ''
        else:
            return x


    # Количество выявленных ошибок
    errors_count = 0
    warnings_count = 0

    for fio in range(2, len(data)):
        p1 = ['Аттестационный период 1']
        p2 = ['Аттестационный период 2']
        if periods == 3:
            p3 = ['Аттестационный период 3']
        god = ['Год']
        predmeti_p1 = ['']

        for i in range(len(data[1])):
            if data[1][i] == 'Аттестационный период 1':
                predmeti_p1.append(data[0][i])
                p1.append(markToInt(data[fio][i]))
            if data[1][i] == 'Аттестационный период 2':
                p2.append(markToInt(data[fio][i]))
            if data[1][i] == 'Аттестационный период 3' and periods == 3:
                p3.append(markToInt(data[fio][i]))
            if data[1][i] == 'Год' and data[0][i] != 'Математика':
                god.append(markToInt(data[fio][i]))


        # Трассировка вывода данных
        # print(data[fio][0])
        # print(p1)
        # print(p2)
        # if periods == 3:
        #      print(p3)
        # print(god)

        out_sheet.append([data[fio][0]])

        out_sheet.append(predmeti_p1)
        out_sheet.append(p1)
        out_sheet.append(p2)
        if periods == 3:
            out_sheet.append(p3)
        out_sheet.append(god)

        # проверка выставления итоговых оценок
        cols = 'ABCDEFGHIJKLMNOPQRSTUVWXWZ'
        for i in range(len(predmeti_p1) - 1):
            if p1[i + 1] != '' or p2[i + 1] != '':
                if periods != 3:
                    a = 0
                    b = 0
                    if p1[i + 1] == '':
                        a == 0
                    elif p1[i + 1] == 'НПА' or p1[i + 1] == 'АЗ':
                        print('НННННННППППППППАААААААА')
                        out_sheet[cols[i + 1] + str(3 + (fio - 2) * 5)].fill = PatternFill('solid', fgColor='FFF273')
                        warnings_count += 1
                        continue
                    else:
                        a = p1[i + 1]

                    if p2[i + 1] == '':
                        b == 0
                    elif p2[i + 1] == 'НПА' or p2[i + 1] == 'АЗ':
                        print('НННННННППППППППАААААААА')
                        out_sheet[cols[i + 1] + str(4 + (fio - 2) * 5)].fill = PatternFill('solid', fgColor='FFF273')
                        warnings_count += 1
                        continue
                    else:
                        b = p2[i + 1]

                    if a == 0 or b == 0:
                        avg = a + b
                    elif p1[i + 1] == 'Зч' or p2[i + 1] == 'Зч':
                        avg = 'Зч'
                    else:
                        avg = ceil((a + b) / 2)

                    # if p1[i + 1] == 'Зч' or p2[i + 1] == 'Зч':
                    #     print('Предмет:',out_sheet[cols[i + 1] + str(2 + (fio - 2) * 5)].value,'| ячейка:', cols[i + 1] + str(5 + (fio - 2) * 5), 'a=',a,'b=',b, '(a+b)/ 2=','Зч','avg=',avg,'god=',god[i+1])
                    # else:
                    #     print('Предмет:',out_sheet[cols[i + 1] + str(2 + (fio - 2) * 5)].value,'| ячейка:', cols[i + 1] + str(5 + (fio - 2) * 5), 'a=',a,'b=',b, '(a+b)/ 2=',(a+b)/2,'avg=',avg,'god=',god[i+1])

                    if avg == god[i + 1]:
                        out_sheet[cols[i + 1] + str(5 + (fio - 2) * 5)].fill = PatternFill('solid', fgColor='A6F16C')
                    else:
                        # Закрашиваем ячейку с ошибкой красным
                        out_sheet[cols[i + 1] + str(5 + (fio - 2) * 5)].fill = PatternFill('solid', fgColor='FF8B73')
                        errors_count += 1
                        # Закрашиваем ФИО ученика с ошибкой красным
                        out_sheet['A' + str(1 + (fio - 2) * 5)].fill = PatternFill('solid', fgColor='FF8B73')

        # переворот названий предметов в ячейке на 90 градусов
        for row in out_sheet['B' + str(2 + (fio - 2) * 5) + ':Z' + str(2 + (fio - 2) * 5)]:
            for cell in row:
                cell.alignment = Alignment(textRotation=90)

    out_sheet.insert_rows(1)
    out_sheet['A1'] = "Количество ошибок:"
    out_sheet['B1'] = errors_count
    if errors_count == 0:
        out_sheet['A1'].fill = PatternFill('solid', fgColor='A6F16C')
        out_sheet['B1'].fill = PatternFill('solid', fgColor='A6F16C')
    else:
        out_sheet['A1'].fill = PatternFill('solid', fgColor='FF8B73')
        out_sheet['B1'].fill = PatternFill('solid', fgColor='FF8B73')

    # выравниевание столбцов по центру
    for letter in 'BCDEFGHIJKLMNOPQRSTUVWXWZ':
        col = out_sheet.column_dimensions[letter]
        col.alignment = Alignment(horizontal='center', vertical='center')

    out_book.save('checked/' + file_name[:-5] + '-ПРОВЕРЕНО(ош=' + str(errors_count) + ',зам='+ str(warnings_count) +').xlsx')